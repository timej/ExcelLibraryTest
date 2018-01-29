import pandas as pd
import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment

def get_list():
    # http://openpyxl.readthedocs.io/en/default/pandas.html#

    path = 'data/sample.txt'
    df = pd.read_csv(path, sep='\t')
    df['注文日'] = pd.to_datetime(df['注文日'])

    wb = Workbook()
    ws = wb.get_active_sheet()

    ws.append([None, None, None, '商品受注リスト'])
    ws.append([None, None, None, None, None, None, '作成日: ' + datetime.date.today().strftime('%Y-%m-%d')])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    begin_row = 3
    end_row = len(df.index) + begin_row

    ws['D1'].font = Font(size=20)
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['G2'].alignment = Alignment(horizontal='right')

    for cell in ws['A:A']:
        cell.number_format = 'YYYY/MM/DD'
    for cell in ws['G:G']:
        cell.number_format = '#,##0'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12 
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # http://openpyxl.readthedocs.io/en/default/worksheet_tables.html

    tab = Table(displayName="Table1", ref="A" + str(begin_row) + ":G" + str(end_row))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save('data/一覧.xlsx')

if __name__ == "__main__":
    get_list()
    
