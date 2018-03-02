import pandas as pd

def pdread(in_path, out_path):
    df = pd.read_excel(in_path, header=None)
    df.to_csv(out_path, sep='\t', header=False, index=False)

def pdwrite(in_path, out_path):
    df = pd.read_csv(in_path, sep='\t', header=None)
    df.to_excel(out_path, header=False, index=False)


def iter_row(row):
    yield [cell.value for cell in row]


def openpyxl_read(in_path, out_path):
    from openpyxl import load_workbook
    import csv

    wb = load_workbook(in_path, read_only=True)
    sheet = wb.worksheets[0]

    with open(out_path, 'w') as f:
        writer = csv.writer(f, delimiter='\t', quoting=csv.QUOTE_MINIMAL)

        for row in sheet.rows:
            writer.writerows(iter_row(row))

'''
# df.to_excelを使うよりwrite_only=Trueにした方が速い
def openpyxl_write(in_path, out_path):
    df = pd.read_csv(in_path, sep='\t', header=None)
    df.to_excel(out_path, header=False, index=False, engine='openpyxl')
'''

def openpyxl_write(in_path, out_path):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    df = pd.read_csv(in_path, sep='\t', header=None)
    
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    wb.save(out_path)


if __name__ == "__main__":
    import sys
    import os
    if not os.path.exists('App_Data/DataOut'):
        os.mkdir('App_Data/DataOut')
    if '-r' in sys.argv:
        if '-x' in sys.argv: 
            pdread('App_Data/DataIn/am0411.xlsx', 'App_Data/DataOut/px.txt')
        elif '-o' in sys.argv:
            openpyxl_read('App_Data/DataIn/am0411.xlsx', 'App_Data/DataOut/po.txt')
        else:
            pdread('App_Data/DataIn/am0411.xls', 'App_Data/DataOut/ps.txt')
    else:
        if '-x' in sys.argv:
            pdwrite('App_Data/DataIn/am0411.txt', 'App_Data/DataOut/px.xlsx')
        elif '-o' in sys.argv:
            openpyxl_write('App_Data/DataIn/am0411.txt', 'App_Data/DataOut/po.xlsx')
        else:
            pdwrite('App_Data/DataIn/am0411.txt', 'App_Data/DataOut/ps.xls')