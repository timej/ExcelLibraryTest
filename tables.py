import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# http://openpyxl.readthedocs.io/en/default/pandas.html#

path = 'App_Data/DataIn/population.txt'
df = pd.read_csv(path, sep='\t')

wb = Workbook()
ws = wb.get_active_sheet()

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

for i in range(2, len(df.index) + 1):
    ws.cell(row=i, column=3).number_format = '#,##0'
    ws.cell(row=i, column=4).number_format = '#,##0.00'
    ws.cell(row=i, column=5).number_format = '#,##0.0'
    ws.cell(row=i, column=6).number_format = '#,##0'
ws.column_dimensions['B'].width = 20 
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12

# http://openpyxl.readthedocs.io/en/default/worksheet_tables.html

tab = Table(displayName="Table2", ref="A1:F" + str(len(df.index) + 1))

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)

wb.save('App_Data/DataOut/population1.xlsx')